'''Task #1'''
def find_top_20(candidates):
    """
    Returns a list of the top twenty students
    """
    students = ()
    for candidate in candidates:
        name = candidate['name']
        total = candidate['extra_scores']+candidate['scores']['math']+candidate['scores']['russian_language']+candidate['scores']['computer_science']
        profile = candidate['scores']['computer_science']+candidate['scores']['math']
        students += ((total, profile, name),)
    return [i[2] for i in sorted(students, reverse=True)[:20]]

candidates = [{'name': 'Dan', 'scores': {'math': 65, 'russian_language': 87, 'computer_science': 86}, 'extra_scores': 5}, {'name': 'Justin', 'scores': {'math': 30, 'russian_language': 31, 'computer_science': 65}, 'extra_scores': 1}, {'name': 'Christina', 'scores': {'math': 50, 'russian_language': 79, 'computer_science': 47}, 'extra_scores': 3}, {'name': 'Johny', 'scores': {'math': 16, 'russian_language': 85, 'computer_science': 16}, 'extra_scores': 3}, {'name': 'Terry', 'scores': {'math': 24, 'russian_language': 39, 'computer_science': 85}, 'extra_scores': 1}, {'name': 'Brandon', 'scores': {'math': 80, 'russian_language': 46, 'computer_science': 4}, 'extra_scores': 9}, {'name': 'Sana', 'scores': {'math': 84, 'russian_language': 7, 'computer_science': 97}, 'extra_scores': 6}, {'name': 'Barbara', 'scores': {'math': 52, 'russian_language': 50, 'computer_science': 61}, 'extra_scores': 0}, {'name': 'Karen', 'scores': {'math': 44, 'russian_language': 52, 'computer_science': 90}, 'extra_scores': 5}, {'name': 'Catherine', 'scores': {'math': 77, 'russian_language': 7, 'computer_science': 7}, 'extra_scores': 7}, {'name': 'Arlene', 'scores': {'math': 90, 'russian_language': 47, 'computer_science': 20}, 'extra_scores': 6}, {'name': 'Eric', 'scores': {'math': 48, 'russian_language': 6, 'computer_science': 11}, 'extra_scores': 8}, {'name': 'Carolsa', 'scores': {'math': 88, 'russian_language': 69, 'computer_science': 65}, 'extra_scores': 7}, {'name': 'Estelle', 'scores': {'math': 39, 'russian_language': 10, 'computer_science': 69}, 'extra_scores': 6}, {'name': 'Amy', 'scores': {'math': 27, 'russian_language': 12, 'computer_science': 21}, 'extra_scores': 2}, {'name': 'Tyler', 'scores': {'math': 88, 'russian_language': 78, 'computer_science': 0}, 'extra_scores': 7}, {'name': 'Tony', 'scores': {'math': 27, 'russian_language': 54, 'computer_science': 41}, 'extra_scores': 6}, {'name': 'Roy', 'scores': {'math': 22, 'russian_language': 5, 'computer_science': 72}, 'extra_scores': 3}, {'name': 'Michael', 'scores': {'math': 7, 'russian_language': 62, 'computer_science': 54}, 'extra_scores': 7}, {'name': 'Angelyn', 'scores': {'math': 44, 'russian_language': 28, 'computer_science': 50}, 'extra_scores': 5}, {'name': 'Jaclyn', 'scores': {'math': 54, 'russian_language': 93, 'computer_science': 0}, 'extra_scores': 7}, {'name': 'Anthony', 'scores': {'math': 4, 'russian_language': 53, 'computer_science': 63}, 'extra_scores': 4}, {'name': 'Monica', 'scores': {'math': 68, 'russian_language': 56, 'computer_science': 2}, 'extra_scores': 0}, {'name': 'Johnbo', 'scores': {'math': 17, 'russian_language': 89, 'computer_science': 50}, 'extra_scores': 1}, {'name': 'Mark', 'scores': {'math': 39, 'russian_language': 30, 'computer_science': 80}, 'extra_scores': 6}, {'name': 'Mary', 'scores': {'math': 36, 'russian_language': 19, 'computer_science': 59}, 'extra_scores': 2}, {'name': 'John', 'scores': {'math': 65, 'russian_language': 36, 'computer_science': 86}, 'extra_scores': 4}, {'name': 'Ethel', 'scores': {'math': 67, 'russian_language': 85, 'computer_science': 59}, 'extra_scores': 0}, {'name': 'Jennifer', 'scores': {'math': 32, 'russian_language': 97, 'computer_science': 27}, 'extra_scores': 8}, {'name': 'Johna', 'scores': {'math': 62, 'russian_language': 55, 'computer_science': 71}, 'extra_scores': 4}, {'name': 'Timothy', 'scores': {'math': 37, 'russian_language': 67, 'computer_science': 89}, 'extra_scores': 4}, {'name': 'Patricia', 'scores': {'math': 44, 'russian_language': 63, 'computer_science': 49}, 'extra_scores': 5}, {'name': 'Alvin', 'scores': {'math': 35, 'russian_language': 81, 'computer_science': 94}, 'extra_scores': 5}, {'name': 'Erica', 'scores': {'math': 49, 'russian_language': 93, 'computer_science': 15}, 'extra_scores': 6}, {'name': 'Hazel', 'scores': {'math': 72, 'russian_language': 79, 'computer_science': 18}, 'extra_scores': 5}, {'name': 'Eleanor', 'scores': {'math': 59, 'russian_language': 46, 'computer_science': 75}, 'extra_scores': 8}, {'name': 'Brenda', 'scores': {'math': 11, 'russian_language': 19, 'computer_science': 53}, 'extra_scores': 4}, {'name': 'Micheal', 'scores': {'math': 73, 'russian_language': 21, 'computer_science': 11}, 'extra_scores': 7}, {'name': 'Dennis', 'scores': {'math': 95, 'russian_language': 69, 'computer_science': 33}, 'extra_scores': 6}, {'name': 'Juan', 'scores': {'math': 33, 'russian_language': 46, 'computer_science': 86}, 'extra_scores': 0}, {'name': 'Eliseo', 'scores': {'math': 79, 'russian_language': 56, 'computer_science': 20}, 'extra_scores': 10}, {'name': 'Grace', 'scores': {'math': 14, 'russian_language': 86, 'computer_science': 29}, 'extra_scores': 1}, {'name': 'James', 'scores': {'math': 79, 'russian_language': 52, 'computer_science': 23}, 'extra_scores': 8}, {'name': 'Salvador', 'scores': {'math': 35, 'russian_language': 25, 'computer_science': 96}, 'extra_scores': 6}, {'name': 'Helen', 'scores': {'math': 69, 'russian_language': 1, 'computer_science': 9}, 'extra_scores': 1}, {'name': 'Adell', 'scores': {'math': 69, 'russian_language': 16, 'computer_science': 91}, 'extra_scores': 8}, {'name': 'Nancy', 'scores': {'math': 82, 'russian_language': 51, 'computer_science': 59}, 'extra_scores': 2}, {'name': 'Robert', 'scores': {'math': 11, 'russian_language': 67, 'computer_science': 41}, 'extra_scores': 3}, {'name': 'Joseph', 'scores': {'math': 93, 'russian_language': 9, 'computer_science': 80}, 'extra_scores': 1}, {'name': 'Chad', 'scores': {'math': 89, 'russian_language': 15, 'computer_science': 35}, 'extra_scores': 3}, {'name': 'Roberta', 'scores': {'math': 96, 'russian_language': 53, 'computer_science': 0}, 'extra_scores': 7}, {'name': 'Edward', 'scores': {'math': 11, 'russian_language': 41, 'computer_science': 91}, 'extra_scores': 3}, {'name': 'Jeremy', 'scores': {'math': 49, 'russian_language': 47, 'computer_science': 48}, 'extra_scores': 9}, {'name': 'Mabel', 'scores': {'math': 71, 'russian_language': 21, 'computer_science': 50}, 'extra_scores': 0}, {'name': 'Richard', 'scores': {'math': 49, 'russian_language': 46, 'computer_science': 35}, 'extra_scores': 5}, {'name': 'Melissa', 'scores': {'math': 69, 'russian_language': 14, 'computer_science': 52}, 'extra_scores': 6}, {'name': 'Roland', 'scores': {'math': 74, 'russian_language': 79, 'computer_science': 96}, 'extra_scores': 5}, {'name': 'Michaela', 'scores': {'math': 38, 'russian_language': 34, 'computer_science': 67}, 'extra_scores': 5}, {'name': 'Rebecca', 'scores': {'math': 1, 'russian_language': 40, 'computer_science': 96}, 'extra_scores': 3}, {'name': 'Michaelo', 'scores': {'math': 3, 'russian_language': 60, 'computer_science': 79}, 'extra_scores': 4}, {'name': 'Joshua', 'scores': {'math': 70, 'russian_language': 39, 'computer_science': 50}, 'extra_scores': 2}, {'name': 'Alfred', 'scores': {'math': 85, 'russian_language': 11, 'computer_science': 16}, 'extra_scores': 10}, {'name': 'Belinda', 'scores': {'math': 53, 'russian_language': 4, 'computer_science': 29}, 'extra_scores': 1}, {'name': 'Ann', 'scores': {'math': 6, 'russian_language': 45, 'computer_science': 34}, 'extra_scores': 6}, {'name': 'Willie', 'scores': {'math': 11, 'russian_language': 24, 'computer_science': 9}, 'extra_scores': 9}, {'name': 'Roberte', 'scores': {'math': 75, 'russian_language': 80, 'computer_science': 60}, 'extra_scores': 5}, {'name': 'Barbaraessa', 'scores': {'math': 97, 'russian_language': 28, 'computer_science': 12}, 'extra_scores': 8}, {'name': 'Rebe', 'scores': {'math': 7, 'russian_language': 43, 'computer_science': 31}, 'extra_scores': 0}, {'name': 'Sean', 'scores': {'math': 67, 'russian_language': 72, 'computer_science': 38}, 'extra_scores': 0}, {'name': 'Brian', 'scores': {'math': 45, 'russian_language': 96, 'computer_science': 61}, 'extra_scores': 9}, {'name': 'Andrew', 'scores': {'math': 14, 'russian_language': 58, 'computer_science': 97}, 'extra_scores': 7}, {'name': 'Irma', 'scores': {'math': 13, 'russian_language': 34, 'computer_science': 53}, 'extra_scores': 6}, {'name': 'Ashley', 'scores': {'math': 69, 'russian_language': 8, 'computer_science': 0}, 'extra_scores': 4}, {'name': 'Roberty', 'scores': {'math': 93, 'russian_language': 81, 'computer_science': 88}, 'extra_scores': 4}, {'name': 'Jessica', 'scores': {'math': 48, 'russian_language': 65, 'computer_science': 11}, 'extra_scores': 7}, {'name': 'Anna', 'scores': {'math': 3, 'russian_language': 99, 'computer_science': 6}, 'extra_scores': 2}, {'name': 'Alyson', 'scores': {'math': 45, 'russian_language': 15, 'computer_science': 70}, 'extra_scores': 7}, {'name': 'Frances', 'scores': {'math': 64, 'russian_language': 43, 'computer_science': 80}, 'extra_scores': 8}, {'name': 'Annauetta', 'scores': {'math': 32, 'russian_language': 2, 'computer_science': 4}, 'extra_scores': 7}, {'name': 'Carolina', 'scores': {'math': 93, 'russian_language': 32, 'computer_science': 0}, 'extra_scores': 4}, {'name': 'Theodore', 'scores': {'math': 26, 'russian_language': 67, 'computer_science': 14}, 'extra_scores': 9}, {'name': 'Deidre', 'scores': {'math': 2, 'russian_language': 27, 'computer_science': 81}, 'extra_scores': 10}, {'name': 'Ernest', 'scores': {'math': 78, 'russian_language': 86, 'computer_science': 47}, 'extra_scores': 0}, {'name': 'Annana', 'scores': {'math': 5, 'russian_language': 11, 'computer_science': 3}, 'extra_scores': 0}, {'name': 'Jesse', 'scores': {'math': 92, 'russian_language': 11, 'computer_science': 37}, 'extra_scores': 6}, {'name': 'Marvin', 'scores': {'math': 78, 'russian_language': 61, 'computer_science': 38}, 'extra_scores': 7}, {'name': 'Nina', 'scores': {'math': 43, 'russian_language': 3, 'computer_science': 31}, 'extra_scores': 9}, {'name': 'Johnka', 'scores': {'math': 96, 'russian_language': 90, 'computer_science': 6}, 'extra_scores': 2}, {'name': 'Jonathan', 'scores': {'math': 29, 'russian_language': 50, 'computer_science': 72}, 'extra_scores': 1}, {'name': 'Blaine', 'scores': {'math': 39, 'russian_language': 26, 'computer_science': 16}, 'extra_scores': 1}, {'name': 'Elaine', 'scores': {'math': 41, 'russian_language': 86, 'computer_science': 27}, 'extra_scores': 10}, {'name': 'Roya', 'scores': {'math': 68, 'russian_language': 100, 'computer_science': 97}, 'extra_scores': 10}, {'name': 'Cliff', 'scores': {'math': 42, 'russian_language': 35, 'computer_science': 10}, 'extra_scores': 2}, {'name': 'Georgia', 'scores': {'math': 21, 'russian_language': 5, 'computer_science': 56}, 'extra_scores': 10}, {'name': 'Joe', 'scores': {'math': 48, 'russian_language': 5, 'computer_science': 38}, 'extra_scores': 10}, {'name': 'Adam', 'scores': {'math': 100, 'russian_language': 71, 'computer_science': 48}, 'extra_scores': 3}, {'name': 'Ana', 'scores': {'math': 45, 'russian_language': 54, 'computer_science': 68}, 'extra_scores': 4}, {'name': 'Jim', 'scores': {'math': 14, 'russian_language': 50, 'computer_science': 48}, 'extra_scores': 9}, {'name': 'Jamesy', 'scores': {'math': 44, 'russian_language': 31, 'computer_science': 37}, 'extra_scores': 10}, {'name': 'Carol', 'scores': {'math': 37, 'russian_language': 66, 'computer_science': 6}, 'extra_scores': 4}]
print(find_top_20(candidates))

'''
Task #2
Функция get_inductees принимает 3 списка одинаковой длины. В первом списке (names) — имена студентов, позволяющие
их точно идентифицировать. Во втором (birthday_years) — год рождения. В третьем (genders) — пол студента.Частично
они отсутствуют ввиду испорченного листа бумаги. Функция возвращает список с именами студентов мужского пола,
которые достигли или могут достигнуть 18 лет в 2021 году, но при этом не старше 30 лет. Cтуденты, по которым
невозможно точно установить - попадают они в список или нет (из-за порчи данных), должны быть выведены отдельно.
'''
def get_inductees(n_list, y_list, s_list):
    persons = list(zip(n_list, y_list, s_list))
    definitely = []
    maybe = []
    for i in persons:
        if i[2] != 'Female' and i[1] is None:
            maybe.append(i[0])
        elif i[2] is None and 18 <= (2021 - i[1]) < 30:
            maybe.append(i[0])
        elif i[2] == 'Male' and 18 <= (2021 - i[1]) < 30:
            definitely.append(i[0])
    print(f'{maybe=}')
    return definitely


names = ["Vasya", "Alice", "Petya", "Jenny", "Fedya", "Viola", "Mark", "Chris", "Margo", 'John', 'Mary', 'Bob', 'Tom', 'Jack']
birthday_years = [1962, 1995, 2000, None, None, None, None, 1998, 2003, None, 2005, 1980, 2002, 1995]
genders = ["Male", "Female", "Male", "Female", "Male", None, None, None, None, 'Male', "Female", "Male", "Male", "Male"]


print(get_inductees(names, birthday_years, genders))

'''
Task 3
Функция find_athlets принимает 3 списка с именами студентов. В первом списке (know_english) — список тех, кто
хорошо владеет английским языком. Второй (sportsmen) — содержит имена тех, кто увлекается спортом. Ну и третий
(more_than_20_years) — предоставляет информацию о тех, кто старше 20 лет. Функция возвращает список имен
студентов, которые подходят под все три критерия.

'''


def find_athlets(one, two, three):
    return [x for x in one if x in two and x in three]


know_english = ["Vasya", "Jimmy", "Max", "Peter", "Eric", "Zoi", "Felix"]
sportsmen = ["Don", "Peter", "Eric", "Jimmy", "Mark"]
more_than_20_years = ["Peter", "Julie", "Jimmy", "Mark", "Max"]
print(find_athlets(know_english, sportsmen, more_than_20_years))

'''
Task 4
Функция make_report_about_top3 принимает словарь (students_avg_scores), в котором ключами являются имена студентов,
а значениями — средний балл за всю учебу. Функция находит ТОП-3 студентов, чей средний балл выше, чем у других.
Функция возвращает ссылку на эксель-файл, в котором упомянуты 3 студента и который потом будет передан в бухгалтерию.
Сам файл необходимо создать внутри функции. Важно сохранить совместимость с Excel, чтобы Ольга Петровна смогла без
проблем получить всю нужную информацию.

'''

import openpyxl
import os


def make_report_about_top3(top):
    top = list(top.items())
    top.sort(key=lambda x: x[1], reverse=True)
    with open('top3.xlsx', 'w') as f:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Топ-3'
        for i in range(len(top[:3])):
            ws.cell(row=i+1, column=1).value = top[i][0]
            ws.cell(row=i+1, column=2).value = top[i][1]
        wb.save('top3.xlsx')
    return os.path.abspath('top3.xlsx')







students_avg_scores = {'Max': 4.964, 'Eric': 4.962, 'Peter': 4.923, 'Mark': 4.957, 'Julie': 4.95, 'Jimmy': 4.973, 'Felix': 4.937, 'Vasya': 4.911, 'Don': 4.936, 'Zoi': 4.937}
print(make_report_about_top3(students_avg_scores))